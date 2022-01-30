import findspark
findspark.init()

from pyspark.sql import SparkSession
from openpyxl.utils import get_column_letter
import openpyxl
import os


# https://stackoverflow.com/a/66599619
def auto_format_cell_width(ws):
    for letter in range(1, ws.max_column):
        maximum_value = 0
        for cell in ws[get_column_letter(letter)]:
            val_to_check = len(str(cell.value))
            if val_to_check > maximum_value:
                maximum_value = val_to_check
        ws.column_dimensions[get_column_letter(letter)].width = maximum_value + 2


spark = SparkSession.builder.appName("Spark_Excel_Write") \
    .config("spark.jars.packages", "com.crealytics:spark-excel_2.12:0.13.7") \
    .getOrCreate()
spark.sparkContext.setLogLevel("ERROR")

inputfilelist = ["C:\\Users\\Vicky\\Minnie\\homes.csv", "C:\\Users\\Vicky\\Minnie\\sample_data.csv"]
outputfile = "C:\\Users\\Vicky\\Desktop\\output.xlsx"

first_time_file_flag = True
for inputfile in inputfilelist:
    sparkDF = spark.read.options(header='true', inferSchema='true').csv(inputfile)
    sheet_name = os.path.splitext(os.path.basename(inputfile))[0]

    # Write sparkDF to output excel
    sparkDF.write.format("com.crealytics.spark.excel") \
        .options(header='true', dataAddress="'" + sheet_name + "'!A3") \
        .mode("append") \
        .save(outputfile)

    # Add text in 1st row using openpyxl and autosize columns
    workbook = openpyxl.load_workbook(outputfile)
    worksheet = workbook[sheet_name]
    worksheet.cell(column=1, row=1, value=sheet_name + " file data:")
    auto_format_cell_width(worksheet)
    workbook.save(outputfile)

    # Remove crc file to prevent spark checksum errors
    crc_file = os.path.join(os.path.dirname(outputfile), "." + os.path.basename(outputfile) + ".crc")
    if os.path.exists(crc_file):
        os.remove(crc_file)

    print("Done: " + inputfile)

spark.stop()
