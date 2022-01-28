import findspark
findspark.init()

from pyspark.sql import SparkSession
from openpyxl.utils import get_column_letter
import pandas as pd
import os


# https://stackoverflow.com/questions/66531396/export-pandas-dataframe-to-xlsx-dealing-with-the-openpyxl-issue-on-python-3-9
def auto_format_cell_width(ws):
    for letter in range(1, ws.max_column):
        maximum_value = 0
        for cell in ws[get_column_letter(letter)]:
            val_to_check = len(str(cell.value))
            if val_to_check > maximum_value:
                maximum_value = val_to_check
        ws.column_dimensions[get_column_letter(letter)].width = maximum_value + 2

spark = SparkSession.builder.appName("Spark_Pandas_Excel") \
    .getOrCreate()
spark.sparkContext.setLogLevel("ERROR")

inputfilelist = ["C:\\Users\\Vicky\\Minnie\\homes.csv", "C:\\Users\\Vicky\\Minnie\\sample_data.csv"]
outputfile = "C:\\Users\\Vicky\\Desktop\\output.xlsx"

first_time_file_flag = True
for inputfile in inputfilelist:
    sparkDF = spark.read.options(header='true', inferSchema='true').csv(inputfile)
    sheet_name = os.path.splitext(os.path.basename(inputfile))[0]
    if first_time_file_flag:
        excel = pd.ExcelWriter(outputfile, engine='openpyxl', mode='w')
        first_time_file_flag = False
    else:
        excel = pd.ExcelWriter(outputfile, engine='openpyxl', mode='a')
    sparkDF.toPandas().to_excel(excel, sheet_name=sheet_name, index=False, startrow=2)
    wsheet = excel.book[sheet_name]
    wsheet.cell(column=1, row=1, value=sheet_name + " file data:")
    auto_format_cell_width(wsheet)
    excel.save()
    print("Done: " + inputfile)

spark.stop()
